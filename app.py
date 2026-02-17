import os
import json
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, jsonify
)

# ---------- Pandas / openpyxl for reading users.xlsx ----------
try:
    import pandas as pd
    _use_pandas = True
except ImportError:
    _use_pandas = False

try:
    import openpyxl
    _use_openpyxl = True
except ImportError:
    _use_openpyxl = False

# ---------- Bot service ----------
import bot_service

app = Flask(__name__)
app.secret_key = os.urandom(24)   # secure random secret

USERS_FILE = os.path.join(os.path.dirname(__file__), "users.xlsx")


# ==========================================
# Auth helpers
# ==========================================
def load_users():
    """Load username/password pairs from users.xlsx.
    Returns dict {username: password} or raises RuntimeError."""
    if not os.path.exists(USERS_FILE):
        raise RuntimeError(f"ไม่พบไฟล์ {USERS_FILE}")

    users = {}
    if _use_pandas:
        df = pd.read_excel(USERS_FILE)
        df.columns = [c.strip().lower() for c in df.columns]
        for _, row in df.iterrows():
            u = str(row.get("username", "")).strip()
            p = str(row.get("password", "")).strip()
            if u:
                users[u] = p
    elif _use_openpyxl:
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        u_col = headers.index("username") if "username" in headers else 0
        p_col = headers.index("password") if "password" in headers else 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            u = str(row[u_col]).strip() if row[u_col] else ""
            p = str(row[p_col]).strip() if row[p_col] else ""
            if u:
                users[u] = p
    else:
        raise RuntimeError("ต้องติดตั้ง pandas หรือ openpyxl ก่อน")
    return users


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ==========================================
# Routes
# ==========================================

@app.route("/", methods=["GET"])
def index():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        try:
            users = load_users()
            if username in users and users[username] == password:
                session["user"] = username
                flash(f"ยินดีต้อนรับ, {username}! 🎉", "success")
                return redirect(url_for("dashboard"))
            else:
                flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง", "error")
        except RuntimeError as e:
            flash(str(e), "error")
        except Exception as e:
            flash(f"เกิดข้อผิดพลาด: {e}", "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("ออกจากระบบแล้ว", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    status = bot_service.get_status()
    return render_template("dashboard.html", user=session["user"], status=status)


# ==========================================
# Bot API endpoints
# ==========================================

@app.route("/api/bot/start", methods=["POST"])
@login_required
def api_bot_start():
    try:
        data = request.get_json(force=True)

        # --- Parse weight_map from form string ---
        weight_map_raw = data.get("weight_map_raw", "")
        weight_map = {}
        if weight_map_raw.strip():
            for line in weight_map_raw.strip().splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split(":")
                if len(parts) >= 2:
                    try:
                        count = int(parts[0].strip())
                        weights = [float(x) for x in parts[1:]]
                        total = sum(weights)
                        if total > 0:
                            weight_map[count] = [w / total for w in weights]
                    except:
                        pass

        # --- Parse forbidden words ---
        forbidden_raw = data.get("forbidden_words", "")
        forbidden_words = []
        if forbidden_raw.strip():
            for w in forbidden_raw.split("+"):
                w = w.strip()
                if w:
                    clean = "".join(w.split())
                    forbidden_words.append(clean)
                    forbidden_words.append(clean.lower())
            forbidden_words = list(set(forbidden_words))

        # --- Parse manual pages ---
        manual_pages_raw = data.get("manual_pages", "")
        manual_pages = []
        if manual_pages_raw.strip():
            for p in manual_pages_raw.split("+"):
                p = p.strip()
                if p.isdigit():
                    manual_pages.append(int(p))

        config = {
            "url": data.get("url", "").strip(),
            "mode": data.get("mode", "random"),
            "speed_mode": data.get("speed_mode", "fast"),
            "rounds": int(data.get("rounds", 1)),
            "forbidden_words": forbidden_words,
            "weight_map": weight_map,
            "dist_mode": int(data.get("dist_mode", 1)),
            "manual_pages": manual_pages,
            "auto_start_first_round": data.get("auto_start_first_round", True),
            "headless": data.get("headless", True),
            "text_pool_raw": data.get("text_pool_raw", ""),
        }

        if not config["url"]:
            return jsonify({"ok": False, "error": "กรุณาใส่ URL"}), 400

        err = bot_service.start_bot(config)
        if err:
            return jsonify({"ok": False, "error": err}), 400

        return jsonify({"ok": True, "message": "บอทเริ่มทำงานแล้ว"})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot/stop", methods=["POST"])
@login_required
def api_bot_stop():
    bot_service.stop_bot()
    return jsonify({"ok": True, "message": "ส่งคำสั่งหยุดแล้ว"})


@app.route("/api/bot/status")
@login_required
def api_bot_status():
    return jsonify(bot_service.get_status())


@app.route("/api/bot/answer", methods=["POST"])
@login_required
def api_bot_answer():
    """รับคำตอบข้อเขียนจาก user แล้ว resume bot"""
    try:
        data = request.get_json(force=True)
        # data["answers"] = { "q_t": { "choices": [...], "is_linked": bool } }
        answers = data.get("answers", {})
        bot_service.submit_text_answers(answers)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot/logs")
@login_required
def api_bot_logs():
    offset = int(request.args.get("offset", 0))
    logs = bot_service.bot_status.get("logs", [])
    return jsonify({
        "logs": logs[offset:],
        "total": len(logs),
        "running": bot_service.bot_status.get("running", False),
        "completed": bot_service.bot_status.get("completed", 0),
        "total_rounds": bot_service.bot_status.get("total", 0),
        "paused": bot_service.bot_status.get("paused", False),
        "pending_inputs": bot_service.bot_status.get("pending_inputs", []),
    })


# ==========================================
# Run
# ==========================================
if __name__ == "__main__":
    print("\n🚀 Starting Google Form Bot Web Server")
    print("📍 http://127.0.0.1:5000\n")
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)